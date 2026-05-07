#!/usr/bin/env python3
"""
import_taste_history.py

Reads the user's 'Screen list.xlsx' history, queries TMDB for metadata,
applies sentiment weights based on personal notes, and exports a pre-calculated
Taste Engine vector (my_taste.json).

v2: Added director-based auteur layer — no extra API calls needed,
    director is read directly from the Excel's column 7.
"""


import json
import sqlite3
import httpx
import time
import os
from collections import defaultdict
from config import TMDB_API_KEY

EXCEL_PATH = '/Users/francescozaccaria/Downloads/Screen list.xlsx'
OUTPUT_PATH = 'site/data/my_taste.json'
DB_PATH = 'data/tmdb_cache.db'

# ── Genre → signal mapping ────────────────────────────────────────────────────
GENRE_SIGNAL_MAP = {
    'Action': ['action', 'mainstream'],
    'Adventure': ['action', 'mainstream'],
    'Animation': ['animation'],
    'Comedy': ['comedy'],
    'Crime': ['crime', 'thriller'],
    'Documentary': ['documentary', 'arthouse'],
    'Drama': ['drama'],
    'Family': ['mainstream'],
    'Fantasy': ['scifi', 'mainstream'],
    'History': ['classic', 'drama'],
    'Horror': ['horror'],
    'Music': ['indie'],
    'Mystery': ['thriller', 'crime'],
    'Romance': ['romance'],
    'Science Fiction': ['scifi', 'visual'],
    'Sci-Fi & Fantasy': ['scifi', 'visual'],
    'TV Movie': ['mainstream'],
    'Thriller': ['thriller'],
    'War': ['drama', 'classic'],
    'Western': ['classic'],
}

# ── Director auteur map ───────────────────────────────────────────────────────
# Each entry: partial lowercase name → list of signals to boost
# Weight applied separately (auteur_weight multiplier below)
AUTEUR_MAP = {
    # Arthouse / Surrealist
    'lanthimos':    ['arthouse', 'visual', 'intense'],
    'lynch':        ['arthouse', 'visual', 'surreal'],
    'cronenberg':   ['arthouse', 'intense', 'horror'],
    'aronofsky':    ['arthouse', 'intense', 'drama'],
    'haneke':       ['arthouse', 'intense', 'european'],
    'tarkovski':    ['arthouse', 'visual', 'classic', 'european'],
    'tarkovsky':    ['arthouse', 'visual', 'classic', 'european'],
    'visconti':     ['arthouse', 'classic', 'european', 'drama'],
    'bergman':      ['arthouse', 'classic', 'european', 'drama'],
    'pasolini':     ['arthouse', 'classic', 'european'],
    'zulawski':     ['arthouse', 'intense', 'european'],
    'kubrick':      ['arthouse', 'visual', 'classic'],
    'cassavetes':   ['arthouse', 'indie'],
    'dreyer':       ['arthouse', 'classic', 'european'],

    # French / European auteurs
    'godard':       ['arthouse', 'european', 'classic'],
    'truffaut':     ['arthouse', 'european', 'classic'],
    'varda':        ['arthouse', 'european', 'indie'],
    'dolan':        ['arthouse', 'european', 'romance'],
    'sciamma':      ['arthouse', 'european', 'romance'],
    'ozon':         ['arthouse', 'european'],
    'denis':        ['arthouse', 'european'],
    'kaurismäki':   ['arthouse', 'european', 'indie'],
    'kaurismaeki':  ['arthouse', 'european', 'indie'],
    'enyedi':       ['arthouse', 'european'],
    'östlund':      ['arthouse', 'european'],
    'ostlund':      ['arthouse', 'european'],
    'trier':        ['arthouse', 'european', 'visual'],   # J. Trier (Thelma)
    'von trier':    ['arthouse', 'european', 'intense'],
    'bellocchio':   ['arthouse', 'european', 'classic'],
    'lang':         ['arthouse', 'european'],
    'benchetrit':   ['arthouse', 'european'],
    'mandico':      ['arthouse', 'european', 'surreal'],
    'hansen-løve':  ['arthouse', 'european'],
    'bonello':      ['arthouse', 'european'],

    # International auteurs
    'chan-wook':    ['arthouse', 'international', 'visual', 'thriller'],
    'bong':         ['arthouse', 'international', 'social'],
    'wong kar':     ['arthouse', 'international', 'visual', 'romance'],
    'kore-eda':     ['arthouse', 'international', 'drama'],
    'joon-ho':      ['arthouse', 'international', 'social'],
    'kurosawa':     ['arthouse', 'international', 'classic'],
    'miyazaki':     ['animation', 'visual', 'arthouse'],
    'mendonca':     ['arthouse', 'international'],
    'hamaguchi':    ['arthouse', 'international', 'drama'],
    'loach':        ['arthouse', 'european', 'social'],
    'dardenne':     ['arthouse', 'european', 'social'],
    'fassbinder':   ['arthouse', 'european', 'classic'],
    'wenders':      ['arthouse', 'european', 'visual'],
    'diop':         ['arthouse', 'international'],
    'sembène':      ['arthouse', 'international'],

    # Visual / Aesthetic auteurs
    'villeneuve':   ['visual', 'scifi', 'arthouse'],
    'nolan':        ['visual', 'mainstream', 'scifi'],
    'fincher':      ['visual', 'thriller'],
    'anderson':     ['visual', 'indie', 'comedy'],  # Wes Anderson
    'del toro':     ['visual', 'horror', 'arthouse'],
    'gondry':       ['visual', 'indie', 'surreal'],
    'kaufman':      ['visual', 'indie', 'arthouse'],
    'refn':         ['visual', 'thriller'],
    'chazelle':     ['visual', 'drama'],
    'lubezki':      ['visual'],
    'jonze':        ['indie', 'visual', 'arthouse'],
    'coppola':      ['indie', 'visual'],  # Sofia
    'mcqueen':      ['arthouse', 'drama', 'social'],  # Steve McQueen

    # Indie / Social realism
    'baker':        ['indie', 'social', 'drama'],
    'gerwig':       ['indie', 'comedy', 'drama'],
    'jarmusch':     ['indie', 'arthouse', 'classic'],
    'korine':       ['indie', 'arthouse'],
    'lonergan':     ['indie', 'drama'],
    'kelly':        ['indie', 'scifi'],
    'reichardt':    ['indie', 'social'],

    # Italians
    'mainetti':     ['indie', 'european', 'drama'],
    'garrone':      ['european', 'drama', 'arthouse'],
    'sorrentino':   ['european', 'drama', 'visual'],
    'moretti':      ['european', 'indie', 'comedy'],
    'guadagnino':   ['european', 'drama', 'visual', 'romance'],
    'cuarón':       ['arthouse', 'visual', 'drama'],
    'cuaron':       ['arthouse', 'visual', 'drama'],

    # Queer cinema
    'campion':      ['arthouse', 'drama', 'romance'],
    'haynes':       ['arthouse', 'drama', 'romance'],
    'almodóvar':    ['arthouse', 'european', 'romance'],
    'almodovar':    ['arthouse', 'european', 'romance'],
}

# Multiplier for auteur signal boost (relative to genre weight = 0.1)
# Set higher to give director more influence than TMDB genre alone
AUTEUR_WEIGHT_MULTIPLIER = 1.5


def init_cache():
    os.makedirs('data', exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS search_cache
                 (title TEXT PRIMARY KEY, result_json TEXT)''')
    conn.commit()
    return conn


def search_tmdb(conn, title, is_series):
    title_str = str(title)
    title_clean = title_str.lower().strip()
    c = conn.cursor()
    c.execute('SELECT result_json FROM search_cache WHERE title = ?', (title_clean,))
    row = c.fetchone()
    if row:
        return json.loads(row[0])

    endpoint = "search/tv" if is_series else "search/movie"
    url = f"https://api.themoviedb.org/3/{endpoint}"
    params = {'api_key': TMDB_API_KEY, 'query': title_str, 'language': 'en-US', 'page': 1}

    try:
        resp = httpx.get(url, params=params, timeout=5.0)
        resp.raise_for_status()
        results = resp.json().get('results', [])
        best_match = results[0] if results else None
        c.execute('INSERT OR REPLACE INTO search_cache VALUES (?, ?)',
                  (title_clean, json.dumps(best_match) if best_match else 'null'))
        conn.commit()
        time.sleep(0.05)
        return best_match
    except Exception as e:
        print(f"  ⚠ TMDB error for '{title}': {e}")
        return None


def fetch_genres():
    genre_map = {}
    for endpoint in ['movie', 'tv']:
        try:
            resp = httpx.get(
                f"https://api.themoviedb.org/3/genre/{endpoint}/list",
                params={'api_key': TMDB_API_KEY}, timeout=5.0
            )
            for g in resp.json().get('genres', []):
                genre_map[g['id']] = g['name']
        except Exception as e:
            print(f"Failed to fetch {endpoint} genres: {e}")
    return genre_map


def apply_auteur_signals(director_str, weight, signals):
    """Apply auteur boost based on director name. Returns list of matched keys."""
    if not director_str:
        return []
    d = director_str.lower()
    matched = []
    for key, auteur_signals in AUTEUR_MAP.items():
        if key in d:
            boost = weight * 0.1 * AUTEUR_WEIGHT_MULTIPLIER
            for s in auteur_signals:
                signals[s] += boost
            matched.append(key)
    return matched


def clamp(val, min_val, max_val):
    return max(min_val, min(max_val, val))


def main():
    import openpyxl
    print(f"Reading {EXCEL_PATH}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print("❌ Excel file not found!")
        return

    conn = init_cache()
    genre_map = fetch_genres()

    items = []

    # ── Main sheet ────────────────────────────────────────────
    ws1 = wb['Streaming Show List']
    for row in ws1.iter_rows(min_row=3, values_only=True):
        title = row[1]
        if not title: continue
        is_series = str(row[5] or '').lower() in ['series', 'tv show']
        director = str(row[7] or '').strip()   # ← column 7
        note = str(row[9] or '').strip().lower()

        weight = 0.5
        if 'nice' in note:        weight = 1.0
        elif 'good' in note:      weight = 1.5
        elif 'bellissimo' in note: weight = 2.0
        elif any(x in note for x in ['boring', 'disappointing', 'bad', 'nope', 'just..bad']):
            weight = -0.5   # soften negatives (still watched it)

        items.append({'title': str(title), 'is_series': is_series,
                      'weight': weight, 'director': director})

    # ── Visti but when ────────────────────────────────────────
    try:
        ws3 = wb['Visti but when']
        for row in ws3.iter_rows(min_row=3, values_only=True):
            title = row[0]
            if not title: continue
            director = str(row[2] or '').strip()
            note = str(row[3] or '').strip().lower()
            weight = 2.0 if 'bellissimo' in note else 1.0
            items.append({'title': str(title), 'is_series': False,
                          'weight': weight, 'director': director})
    except KeyError:
        pass

    print(f"Found {len(items)} items to process.")

    signals = defaultdict(float)
    decades = defaultdict(float)
    total_processed = 0
    auteur_hits = defaultdict(int)   # for stats
    auteur_details = []              # for verbose report

    for idx, item in enumerate(items):
        if idx % 100 == 0:
            print(f"  {idx}/{len(items)}...")

        match = search_tmdb(conn, item['title'], item['is_series'])
        if not match:
            match = search_tmdb(conn, item['title'], not item['is_series'])
        if not match:
            continue

        total_processed += 1
        normalized = item['weight'] * 0.1

        # ── Genre signals ─────────────────────────────────────
        g_ids = match.get('genre_ids', [])
        resolved_genres = [genre_map.get(gid) for gid in g_ids if gid in genre_map]
        for g in resolved_genres:
            for s in GENRE_SIGNAL_MAP.get(g, []):
                signals[s] += normalized

        # ── Decade ────────────────────────────────────────────
        date_str = match.get('release_date') or match.get('first_air_date', '')
        if date_str and len(date_str) >= 4:
            year = int(date_str[:4])
            decade = f"{year - (year % 10)}s"
            decades[decade] += normalized

        # ── Auteur director boost ─────────────────────────────
        matched_auteurs = apply_auteur_signals(item['director'], item['weight'], signals)
        for a in matched_auteurs:
            auteur_hits[a] += 1
        if matched_auteurs and item['weight'] >= 1.0:
            auteur_details.append({
                'title': item['title'],
                'director': item['director'],
                'weight': item['weight'],
                'auteurs': matched_auteurs,
            })

    # ── Normalize ─────────────────────────────────────────────
    max_sig = max([abs(v) for v in signals.values()] + [1.0])
    for s in signals:
        signals[s] = clamp(signals[s] / max_sig, -1, 1)

    max_dec = max([abs(v) for v in decades.values()] + [1.0])
    for d in decades:
        decades[d] = clamp(decades[d] / max_dec, -1, 1)

    # ── Adventurousness ───────────────────────────────────────
    niche = ['arthouse', 'international', 'european', 'indie', 'auteur', 'surreal']
    niche_avg = sum(max(0, signals.get(s, 0)) for s in niche) / len(niche)
    adventurousness = clamp(niche_avg * 2 + 0.3, 0, 1)

    taste_vector = {
        "signals": dict(signals),
        "decades": dict(decades),
        "adventurousness": round(adventurousness, 3),
        "totalRatings": total_processed,
        "lastUpdated": None,
        "source": "excel_import_v2_auteur",
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(taste_vector, f, indent=2)

    # ── Report ────────────────────────────────────────────────
    print(f"\n✅ Done! Processed {total_processed}/{len(items)} items.")
    print(f"\n── Final Taste Vector ──────────────────────────────────")
    for k, v in sorted(signals.items(), key=lambda x: -x[1]):
        if abs(v) < 0.005: continue
        bar = '█' * int(abs(v) * 25)
        sign = '+' if v >= 0 else '-'
        print(f"  {k:15s}: {sign}{abs(v):.3f}  {bar}")

    print(f"\n── Auteur Director Hits ────────────────────────────────")
    for a, cnt in sorted(auteur_hits.items(), key=lambda x: -x[1]):
        print(f"  {a}: {cnt} film/serie")

    print(f"\n── Top Auteur Contributions (weight ≥ 1.0) ────────────")
    for d in sorted(auteur_details, key=lambda x: -x['weight'])[:20]:
        print(f"  [{d['weight']}] {d['title']} ({d['director']}) → {d['auteurs']}")

    print(f"\n  Adventurousness: {adventurousness:.2f}")
    print(f"  Exported to {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
